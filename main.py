import sys
import webbrowser as wb
from PyQt5 import QtWidgets, QtGui
from PyQt5.QtWidgets import QApplication, QMainWindow, QDialog, QCalendarWidget, QLineEdit,QMessageBox
from openpyxl import load_workbook
##################################################
class MyWindow(QMainWindow):
    def __init__(self):
        super(MyWindow, self).__init__()
        self.initUI()

    def initUI(self):
        self.tytul=QtWidgets.QLabel(self)
        self.tytul.setText("Kalendarz imprez 2022")
        self.tytul.move(260,50)
        czcionka=QtGui.QFont()
        czcionka.setPointSize(20)
        self.tytul.setFont(czcionka)
        self.tytul.adjustSize()

        self.obraz1 = QtWidgets.QLabel(self)
        self.obraz1.setGeometry(50, 100, 300, 200)
        self.obraz1.setPixmap(QtGui.QPixmap("JOTPEGI/domek11.jpg"))
        self.obraz1.setScaledContents(True)

        self.wybor1=QtWidgets.QPushButton(self)
        self.wybor1.setText("Domek 1")
        self.wybor1.move(140,320)
        self.wybor1.clicked.connect(lambda :self.nastepna_strona(1))

        self.lokalizacja1 = QtWidgets.QPushButton(self)
        self.lokalizacja1.setText("Lokalizacja")
        self.lokalizacja1.move(140, 360)
        self.lokalizacja1.clicked.connect(lambda: self.przekieruj("https://www.google.com/maps/place/50.836794,17.446463/data=!4m6!3m5!1s0!7e2!8m2!3d50.836794!4d17.4464634?utm_source=mstt_1&entry=gps"))

        self.obraz2 = QtWidgets.QLabel(self)
        self.obraz2.setGeometry(450, 100, 300, 200)
        self.obraz2.setPixmap(QtGui.QPixmap("JOTPEGI/domek22.jpg"))
        self.obraz2.setScaledContents(True)

        self.wybor2 = QtWidgets.QPushButton(self)
        self.wybor2.setText("Domek 2")
        self.wybor2.move(540,320)
        self.wybor2.clicked.connect(lambda :self.nastepna_strona(2))

        self.lokalizacja2 = QtWidgets.QPushButton(self)
        self.lokalizacja2.setText("Lokalizacja")
        self.lokalizacja2.move(540,360)
        self.lokalizacja2.clicked.connect(lambda: self.przekieruj("https://www.google.com/maps/place/50%C2%B059'59.9%22N+19%C2%B016'12.0%22E/@50.999974,19.2678043,17z/data=!3m1!4b1!4m6!3m5!1s0!7e2!8m2!3d50.9999742!4d19.2699934"))

    def test(self):
        print("Task failed successfully")

    def przekieruj(self,link):
        wb.open(link)

    def nastepna_strona(self,id):
        win.setCurrentIndex(win.currentIndex()+id)

##################################################

class Strona2(QDialog):
    def __init__(self):
        super(Strona2, self).__init__()
        self.initUI()
        self.plik = load_workbook("EXCEL/Terminy_domek.xlsx")
        self.ws = self.plik["Domek_1"]

    def initUI(self):
        self.kalendarz=QCalendarWidget(self)
        self.kalendarz.move(20,20)
        self.kalendarz.setGridVisible(True)

        self.status=QtWidgets.QLabel(self)
        self.status.setText("Status")
        self.status.move(400, 20)

        self.imie = QtWidgets.QLabel(self)
        self.imie.setText("Imię:")
        self.imie.move(400, 50)

        self.imie_dane=QLineEdit(self)
        self.imie_dane.move(400, 70)

        self.nazwisko = QtWidgets.QLabel(self)
        self.nazwisko.setText("Nazwisko:")
        self.nazwisko.move(400, 100)

        self.nazwisko_dane = QLineEdit(self)
        self.nazwisko_dane.move(400, 120)

        self.telefon = QtWidgets.QLabel(self)
        self.telefon.setText("Telefon:")
        self.telefon.move(400, 150)

        self.telefon_dane = QLineEdit(self)
        self.telefon_dane.move(400, 170)

        self.kalendarz.clicked.connect(lambda dzien:self.wolne(dzien))

        self.rez=0
        self.ind=0
        self.termin=""

        self.rezerwuj=QtWidgets.QPushButton(self)
        self.rezerwuj.setText("Rezerwuj")
        self.rezerwuj.move(430, 250)
        self.rezerwuj.clicked.connect(self.zapis)
        self.rezerwuj.hide()

        self.szczegoly=QtWidgets.QPushButton(self)
        self.szczegoly.setText("Szczegóły")
        self.szczegoly.move(80, 250)
        self.szczegoly.clicked.connect(self.info)
        self.szczegoly.hide()

        self.odwolaj = QtWidgets.QPushButton(self)
        self.odwolaj.setText("Odwołaj")
        self.odwolaj.move(170, 250)
        self.odwolaj.clicked.connect(self.usun)
        self.odwolaj.hide()

        self.cofnij=QtWidgets.QPushButton(self)
        self.cofnij.setText("Cofnij")
        self.cofnij.move(50, 430)
        self.cofnij.clicked.connect(lambda :self.poprzednia_strona(1))

    def poprzednia_strona(self,id):
        win.setCurrentIndex(win.currentIndex()-id)

    def wolne(self,x):
        tekst = str(x)
        self.termin=tekst
        for i in range(self.ws["H1"].value):
            if self.termin==self.ws[chr(65)+str(i+2)].value:
                self.status.setText("Status: Zajęte")
                self.status.adjustSize()
                self.rez=1
                self.szczegoly.show()
                self.odwolaj.show()
                self.rezerwuj.hide()
                self.ind=(i+2)
                break
            else:
                self.status.setText("Status: Wolne")
                self.status.adjustSize()
                self.rez=0
                self.szczegoly.hide()
                self.odwolaj.hide()
                self.rezerwuj.show()

    def zapis(self):
        if self.rez==0:
            if self.imie_dane.text()!="" and self.nazwisko_dane.text()!="" and self.telefon_dane.text()!="":
                self.ws.append([self.termin,self.imie_dane.text(),self.nazwisko_dane.text(),self.telefon_dane.text()])
                self.ws["H1"].value=self.ws["H1"].value+1
                self.ind=self.ws["H1"].value
                self.plik.save("EXCEL/Terminy_domek.xlsx")
                pow = QMessageBox()
                pow.setWindowTitle("Informacja")
                pow.setText("Pomyślnie zarezerwoano.")
                pow.setIcon(QMessageBox.Information)
                x = pow.exec_()
            else:
                pow = QMessageBox()
                pow.setWindowTitle("Informacja")
                pow.setText("Brak danych.")
                pow.setIcon(QMessageBox.Information)
                x = pow.exec_()
        else:
            pow = QMessageBox()
            pow.setWindowTitle("Informacja")
            pow.setText("Termin jest już zajęty.")
            pow.setIcon(QMessageBox.Critical)
            x = pow.exec_()

    def usun(self):
        if self.rez==1:
            self.ws.delete_rows(self.ind)
            self.ws["H1"].value = self.ws["H1"].value -1
            self.plik.save("EXCEL/Terminy_domek.xlsx")
            pow = QMessageBox()
            pow.setWindowTitle("Informacja")
            pow.setText("Usunięto rezerwacje.")
            pow.setIcon(QMessageBox.Information)
            x = pow.exec_()
        else:
            pow = QMessageBox()
            pow.setWindowTitle("Informacja")
            pow.setText("Brak rezerwacji do odwołania")
            pow.setIcon(QMessageBox.Critical)
            x = pow.exec_()

    def info(self):
        i=self.ws[chr(66)+str(self.ind)].value
        n=self.ws[chr(67)+str(self.ind)].value
        t=self.ws[chr(68)+str(self.ind)].value
        pow = QMessageBox()
        pow.setWindowTitle("Szczegóły")
        pow.setText("Imię: "+str(i)+"\n"+"Nazwisko: "+str(n)+"\n"+"Telefon: "+str(t))
        pow.setIcon(QMessageBox.Information)
        pow.adjustSize()
        x = pow.exec_()

##################################################

class Strona3(QDialog):
    def __init__(self):
        super(Strona3, self).__init__()
        self.initUI()
        self.plik = load_workbook("EXCEL/Terminy_domek.xlsx")
        self.ws = self.plik["Domek_2"]

    def initUI(self):
        self.kalendarz=QCalendarWidget(self)
        self.kalendarz.move(20,20)
        self.kalendarz.setGridVisible(True)

        self.status=QtWidgets.QLabel(self)
        self.status.setText("Status")
        self.status.move(400, 20)

        self.imie = QtWidgets.QLabel(self)
        self.imie.setText("Imię:")
        self.imie.move(400, 50)

        self.imie_dane=QLineEdit(self)
        self.imie_dane.move(400, 70)

        self.nazwisko = QtWidgets.QLabel(self)
        self.nazwisko.setText("Nazwisko:")
        self.nazwisko.move(400, 100)

        self.nazwisko_dane = QLineEdit(self)
        self.nazwisko_dane.move(400, 120)

        self.telefon = QtWidgets.QLabel(self)
        self.telefon.setText("Telefon:")
        self.telefon.move(400, 150)

        self.telefon_dane = QLineEdit(self)
        self.telefon_dane.move(400, 170)

        self.kalendarz.clicked.connect(lambda dzien:self.wolne(dzien))

        self.rez=0
        self.ind=0
        self.termin=""

        self.rezerwuj=QtWidgets.QPushButton(self)
        self.rezerwuj.setText("Rezerwuj")
        self.rezerwuj.move(430, 250)
        self.rezerwuj.clicked.connect(self.zapis)
        self.rezerwuj.hide()

        self.szczegoly=QtWidgets.QPushButton(self)
        self.szczegoly.setText("Szczegóły")
        self.szczegoly.move(80, 250)
        self.szczegoly.clicked.connect(self.info)
        self.szczegoly.hide()

        self.odwolaj = QtWidgets.QPushButton(self)
        self.odwolaj.setText("Odwołaj")
        self.odwolaj.move(170, 250)
        self.odwolaj.clicked.connect(self.usun)
        self.odwolaj.hide()

        self.cofnij=QtWidgets.QPushButton(self)
        self.cofnij.setText("Cofnij")
        self.cofnij.move(50, 430)
        self.cofnij.clicked.connect(lambda: self.poprzednia_strona(2))

    def poprzednia_strona(self, id):
        win.setCurrentIndex(win.currentIndex() - id)

    def wolne(self,x):
        tekst = str(x)
        self.termin=tekst
        for i in range(self.ws["H1"].value):
            if self.termin==self.ws[chr(65)+str(i+2)].value:
                self.status.setText("Status: Zajęte")
                self.status.adjustSize()
                self.rez=1
                self.szczegoly.show()
                self.odwolaj.show()
                self.rezerwuj.hide()
                self.ind=(i+2)
                break
            else:
                self.status.setText("Status: Wolne")
                self.status.adjustSize()
                self.rez=0
                self.szczegoly.hide()
                self.odwolaj.hide()
                self.rezerwuj.show()

    def zapis(self):
        if self.rez==0:
            if self.imie_dane.text()!="" and self.nazwisko_dane.text()!="" and self.telefon_dane.text()!="":
                self.ws.append([self.termin,self.imie_dane.text(),self.nazwisko_dane.text(),self.telefon_dane.text()])
                self.ws["H1"].value=self.ws["H1"].value+1
                self.ind=self.ws["H1"].value
                self.plik.save("EXCEL/Terminy_domek.xlsx")
                pow = QMessageBox()
                pow.setWindowTitle("Informacja")
                pow.setText("Pomyślnie zarezerwoano.")
                pow.setIcon(QMessageBox.Information)
                x = pow.exec_()
            else:
                pow = QMessageBox()
                pow.setWindowTitle("Informacja")
                pow.setText("Brak danych.")
                pow.setIcon(QMessageBox.Information)
                x = pow.exec_()
        else:
            pow = QMessageBox()
            pow.setWindowTitle("Informacja")
            pow.setText("Termin jest już zajęty.")
            pow.setIcon(QMessageBox.Critical)
            x = pow.exec_()

    def usun(self):
        if self.rez==1:
            self.ws.delete_rows(self.ind)
            self.ws["H1"].value = self.ws["H1"].value -1
            self.plik.save("EXCEL/Terminy_domek.xlsx")
            pow = QMessageBox()
            pow.setWindowTitle("Informacja")
            pow.setText("Usunięto rezerwacje.")
            pow.setIcon(QMessageBox.Information)
            x = pow.exec_()
        else:
            pow = QMessageBox()
            pow.setWindowTitle("Informacja")
            pow.setText("Brak rezerwacji do odwołania")
            pow.setIcon(QMessageBox.Critical)
            x = pow.exec_()

    def info(self):
        i=self.ws[chr(66)+str(self.ind)].value
        n=self.ws[chr(67)+str(self.ind)].value
        t=self.ws[chr(68)+str(self.ind)].value
        pow = QMessageBox()
        pow.setWindowTitle("Szczegóły")
        pow.setText("Imię: "+str(i)+"\n"+"Nazwisko: "+str(n)+"\n"+"Telefon: "+str(t))
        pow.setIcon(QMessageBox.Information)
        pow.adjustSize()
        x = pow.exec_()
##################################################

app=QApplication(sys.argv)
win=QtWidgets.QStackedWidget()
st1=MyWindow()
win.addWidget(st1)
st2=Strona2()
win.addWidget(st2)
st3=Strona3()
win.addWidget(st3)
win.setFixedHeight(500)
win.setFixedWidth(800)
win.setWindowTitle("Terminarz")

win.show()
sys.exit(app.exec_())